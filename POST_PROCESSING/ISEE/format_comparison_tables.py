import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
import os
import CFG_POST_PROCESS_ISEE as cfg
from copy import copy
from openpyxl.styles import PatternFill
import numpy as np

def copy_template_block(template_sheet, consolidated_sheet, start_row):
    """
    Copies the content and formatting of the template sheet to the consolidated sheet.

    Parameters:
    - template_sheet: the source template sheet.
    - consolidated_sheet: the destination sheet to copy to.
    - start_row: the starting row in the destination sheet.
    """
    for row_idx, row in enumerate(template_sheet.iter_rows(), start=start_row):
        for col_idx, cell in enumerate(row, start=1):
            new_cell = consolidated_sheet.cell(row=row_idx, column=col_idx, value=cell.value)
            if cell.has_style:
                new_cell._style = copy(cell._style)  # Copy cell style
            new_cell.number_format = cell.number_format
            new_cell.font = copy(cell.font)
            new_cell.fill = copy(cell.fill)
            new_cell.border = copy(cell.border)
            new_cell.alignment = copy(cell.alignment)

    # Copy merged cells
    for merged_range in template_sheet.merged_cells.ranges:
        consolidated_sheet.merge_cells(start_row=start_row + merged_range.min_row - 1,
                                       end_row=start_row + merged_range.max_row - 1,
                                       start_column=merged_range.min_col,
                                       end_column=merged_range.max_col)

def map_values_to_template(template_path, sheet, df_data, output_path, mapping_rules, value_col_ref, value_col_plan, signif_test_col, ref_plan='GERBL2_2014'):
    """
    Fills the Excel template with data from a CSV file and saves it.

    Parameters:
    - template_path: str, path to the Excel template file.
    - df_data: df, Dataframe containing the data.
    - output_path: str, path to the output Excel file.
    - mapping_rules: dict, rules defining which cells map to which conditions.
    - value_col: Column to write results into.
    """
    # Load the template and CSV data
    template = load_workbook(template_path)
    template_sheet = template[sheet]

    consolidated_workbook = load_workbook(output_path)
    consolidated_sheet = consolidated_workbook.active
    consolidated_sheet.title = "Consolidated Results"

    next_available_row = consolidated_sheet.max_row+1
    n_rows = template_sheet.max_row

    color_red = 'FFFF6666'
    color_green = 'FF90EE90'
    color_white = 'FFFFFFFF'

    # Get unique combinations of PI and Variable
    unique_pis = df_data[['PI_NAME', 'VARIABLE']].drop_duplicates()
    for _, row in unique_pis.iterrows():
        pi = row['PI_NAME']
        variable = row['VARIABLE']
        print(pi, variable)


        # Filter data for the current PI/Variable
        filtered_data = df_data[(df_data['PI_NAME'] == pi) & (df_data['VARIABLE'] == variable)]

        # Copy the template block into the consolidated sheet
        copy_template_block(template_sheet, consolidated_sheet, next_available_row)

        consolidated_sheet.cell(row=next_available_row + 2, column=1, value=pi)
        consolidated_sheet.cell(row=next_available_row + 2, column=2, value=variable)

        # Update template cells based on rules
        for _, data_row in filtered_data.iterrows():
            section = data_row['SECT_NAME']
            plan = data_row['PLAN_NAME']
            if plan == ref_plan:
                value_col = value_col_ref
            else:
                value_col = value_col_plan

            supply_scenario = data_row['SUPPLY_SCEN']
            value = np.round(data_row[value_col], decimals=2)
            direction = data_row[signif_test_col]
            if direction == '-':
                color = color_red
            elif direction == '+':
                color = color_green
            else:
                color = color_white

            # Get the cell based on mapping rules
            # Get the cell based on mapping rules
            list_index = []
            print(supply_scenario, section, plan)
            if (supply_scenario, section, plan) in mapping_rules:
                col_letter, relative_row = mapping_rules[(supply_scenario, section, plan)]
                col_index = ord(col_letter.upper()) - ord('A') + 1
                row_index = next_available_row + relative_row - 1
                print(row_index, col_index)
                list_index.append((row_index, col_index))
                cell = consolidated_sheet.cell(row=row_index, column=col_index, value=value)
                cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")

        next_available_row += n_rows # +2 for spacing between blocks

    # Define the grey fill color
    grey_fill = PatternFill(start_color="FFB0B0B0", end_color="FFB0B0B0", fill_type="solid")

    # Fill empty cells with grey
    for row in consolidated_sheet.iter_rows(min_row=1, max_row=consolidated_sheet.max_row,
                                            min_col=4, max_col=consolidated_sheet.max_column):
        for cell in row:
            if cell.value is None:  # Check if the cell is empty
                cell.fill = grey_fill
    consolidated_workbook.save(output_path)
    #consolidated_workbook.close()

# Example usage
template_path = r"P:\GLAM\Dashboard\ISEE_Dash_portable\results_template.xlsx"
outfolder = os.path.join(cfg.POST_PROCESS_RES, r'PI_CSV_RESULTS_20250212')
csv_path = os.path.join(outfolder, f'PIs_SUMMARY_RESULTS_20250212.csv')
# Define mapping rules (example: map section, plan, supply_scenario to Excel cells)

# mapping_rules_no_country = {
#     ("Historical", "LKO", "GERBL2_2014"): ("E", 3),  # Maps to cell B5
#     ("Historical", "USL_US", "GERBL2_2014"): ("F", 3),  # Maps to cell B5
#     ("Historical", "USL_DS", "GERBL2_2014"): ("G", 3),  # Maps to cell B5
#     ("Historical", "SLR_US", "GERBL2_2014"): ("H", 3),  # Maps to cell B5
#     ("Historical", "SLR_DS", "GERBL2_2014"): ("I", 3),  # Maps to cell B5
#     ("STO330", "LKO", "GERBL2_2014"): ("J", 3),  # Maps to cell B5
#     ("STO330", "USL_US", "GERBL2_2014"): ("K", 3),  # Maps to cell B5
#     ("STO330", "USL_DS", "GERBL2_2014"): ("L", 3),  # Maps to cell B5
#     ("STO330", "SLR_US", "GERBL2_2014"): ("M", 3),  # Maps to cell B5
#     ("STO330", "SLR_DS", "GERBL2_2014"): ("N", 3),  # Maps to cell B5
#     ("RCP45", "LKO", "GERBL2_2014"): ("O", 3),  # Maps to cell B5
#     ("RCP45", "USL_US", "GERBL2_2014"): ("P", 3),  # Maps to cell B5
#     ("RCP45", "USL_DS", "GERBL2_2014"): ("Q", 3),  # Maps to cell B5
#     ("RCP45", "SLR_US", "GERBL2_2014"): ("R", 3),  # Maps to cell B5
#     ("RCP45", "SLR_DS", "GERBL2_2014"): ("S", 3),  # Maps to cell B5
#
#     ("Historical", "LKO", "GERBL2_2014_ComboC"): ("E", 4),
#     ("Historical", "USL_US", "GERBL2_2014_ComboC"): ("F", 4),
#     ("Historical", "USL_DS", "GERBL2_2014_ComboC"): ("G", 4),
#     ("Historical", "SLR_US", "GERBL2_2014_ComboC"): ("H", 4),
#     ("Historical", "SLR_DS", "GERBL2_2014_ComboC"): ("I", 4),
#     ("STO330", "LKO", "GERBL2_2014_ComboC"): ("J", 4),
#     ("STO330", "USL_US", "GERBL2_2014_ComboC"): ("K", 4),
#     ("STO330", "USL_DS", "GERBL2_2014_ComboC"): ("L", 4),
#     ("STO330", "SLR_US", "GERBL2_2014_ComboC"): ("M", 4),
#     ("STO330", "SLR_DS", "GERBL2_2014_ComboC"): ("N", 4),
#     ("RCP45", "LKO", "GERBL2_2014_ComboC"): ("O", 4),
#     ("RCP45", "USL_US", "GERBL2_2014_ComboC"): ("P", 4),
#     ("RCP45", "USL_DS", "GERBL2_2014_ComboC"): ("Q", 4),
#     ("RCP45", "SLR_US", "GERBL2_2014_ComboC"): ("R", 4),
#     ("RCP45", "SLR_DS", "GERBL2_2014_ComboC"): ("S", 4),
#
#     ("Historical", "LKO", "PreProject"): ("E", 5),
#     ("Historical", "USL_US", "PreProject"): ("F", 5),
#     #("Historical", "USL_DS", "PreProject"): ("G", 5),
#     ("Historical", "SLR_US", "PreProject"): ("H", 5),
#     ("Historical", "SLR_DS", "PreProject"): ("I", 5),
#     ("STO330", "LKO", "PreProject"): ("J", 5),
#     ("STO330", "USL_US", "PreProject"): ("K", 5),
#     #("STO330", "USL_DS", "PreProject"): ("L", 5),
#     ("STO330", "SLR_US", "PreProject"): ("M", 5),
#     ("STO330", "SLR_DS", "PreProject"): ("N", 5),
#     ("RCP45", "LKO", "PreProject"): ("O", 5),
#     ("RCP45", "USL_US", "PreProject"): ("P", 5),
#     #("RCP45", "USL_DS", "PreProject"): ("Q", 5),
#     ("RCP45", "SLR_US", "PreProject"): ("R", 5),
#     ("RCP45", "SLR_DS", "PreProject"): ("S", 5),
#
# }
#
# mapping_rules_2 = {
#     ("Historical", "LKO", "GERBL2_2014"): ("E", 3),  # Maps to cell B5
#     ("Historical", "USL_US", "GERBL2_2014"): ("F", 3),  # Maps to cell B5
#     ("Historical", "USL_DS", "GERBL2_2014"): ("G", 3),  # Maps to cell B5
#     ("Historical", "SLR_US", "GERBL2_2014"): ("H", 3),  # Maps to cell B5
#     ("Historical", "SLR_DS", "GERBL2_2014"): ("I", 3),  # Maps to cell B5
#     ("STO330", "LKO", "GERBL2_2014"): ("E", 8),  # Maps to cell B5
#     ("STO330", "USL_US", "GERBL2_2014"): ("F", 8),  # Maps to cell B5
#     ("STO330", "USL_DS", "GERBL2_2014"): ("G", 8),  # Maps to cell B5
#     ("STO330", "SLR_US", "GERBL2_2014"): ("H", 8),  # Maps to cell B5
#     ("STO330", "SLR_DS", "GERBL2_2014"): ("I", 8),  # Maps to cell B5
#     ("RCP45", "LKO", "GERBL2_2014"): ("E", 13),  # Maps to cell B5
#     ("RCP45", "USL_US", "GERBL2_2014"): ("F", 13),  # Maps to cell B5
#     ("RCP45", "USL_DS", "GERBL2_2014"): ("G", 13),  # Maps to cell B5
#     ("RCP45", "SLR_US", "GERBL2_2014"): ("H", 13),  # Maps to cell B5
#     ("RCP45", "SLR_DS", "GERBL2_2014"): ("I", 13),  # Maps to cell B5
#
#     ("Historical", "LKO", "GERBL2_2014_ComboC"): ("E", 4),
#     ("Historical", "USL_US", "GERBL2_2014_ComboC"): ("F", 4),
#     ("Historical", "USL_DS", "GERBL2_2014_ComboC"): ("G", 4),
#     ("Historical", "SLR_US", "GERBL2_2014_ComboC"): ("H", 4),
#     ("Historical", "SLR_DS", "GERBL2_2014_ComboC"): ("I", 4),
#     ("STO330", "LKO", "GERBL2_2014_ComboC"): ("E", 9),
#     ("STO330", "USL_US", "GERBL2_2014_ComboC"): ("F", 9),
#     ("STO330", "USL_DS", "GERBL2_2014_ComboC"): ("G", 9),
#     ("STO330", "SLR_US", "GERBL2_2014_ComboC"): ("H", 9),
#     ("STO330", "SLR_DS", "GERBL2_2014_ComboC"): ("I", 9),
#     ("RCP45", "LKO", "GERBL2_2014_ComboC"): ("E", 14),
#     ("RCP45", "USL_US", "GERBL2_2014_ComboC"): ("F", 14),
#     ("RCP45", "USL_DS", "GERBL2_2014_ComboC"): ("G", 14),
#     ("RCP45", "SLR_US", "GERBL2_2014_ComboC"): ("H", 14),
#     ("RCP45", "SLR_DS", "GERBL2_2014_ComboC"): ("I", 14),
#
#     ("Historical", "LKO", "PreProject"): ("E", 5),
#     ("Historical", "USL_US", "PreProject"): ("F", 5),
#     #("Historical", "USL_DS", "PreProject"): ("G", 5),
#     ("Historical", "SLR_US", "PreProject"): ("H", 5),
#     ("Historical", "SLR_DS", "PreProject"): ("I", 5),
#     ("STO330", "LKO", "PreProject"): ("E", 10),
#     ("STO330", "USL_US", "PreProject"): ("F", 10),
#     #("STO330", "USL_DS", "PreProject"): ("G", 10),
#     ("STO330", "SLR_US", "PreProject"): ("H", 10),
#     ("STO330", "SLR_DS", "PreProject"): ("I", 10),
#     ("RCP45", "LKO", "PreProject"): ("E", 15),
#     ("RCP45", "USL_US", "PreProject"): ("F", 15),
#     #("RCP45", "USL_DS", "PreProject"): ("G", 15),
#     ("RCP45", "SLR_US", "PreProject"): ("H", 15),
#     ("RCP45", "SLR_DS", "PreProject"): ("I", 15),
# }

# mapping_rules_2 = {
#     ("Historical", "LKO", "GERBL2_2014"): ("D", 3),  # Maps to cell B5
#     ("Historical", "USL_US", "GERBL2_2014"): ("E", 3),  # Maps to cell B5
#     ("Historical", "USL_DS", "GERBL2_2014"): ("F", 3),  # Maps to cell B5
#     ("Historical", "SLR_US", "GERBL2_2014"): ("G", 3),  # Maps to cell B5
#     ("Historical", "SLR_DS", "GERBL2_2014"): ("H", 3),  # Maps to cell B5
#     ("RCP45", "LKO", "GERBL2_2014"): ("D", 9),  # Maps to cell B5
#     ("RCP45", "USL_US", "GERBL2_2014"): ("E", 9),  # Maps to cell B5
#     ("RCP45", "USL_DS", "GERBL2_2014"): ("F", 9),  # Maps to cell B5
#     ("RCP45", "SLR_US", "GERBL2_2014"): ("G", 9),  # Maps to cell B5
#     ("RCP45", "SLR_DS", "GERBL2_2014"): ("H", 9),  # Maps to cell B5
#
#     ("Historical", "LKO", "GERBL2_2014_ComboC"): ("D", 4),
#     ("Historical", "USL_US", "GERBL2_2014_ComboC"): ("E", 4),
#     ("Historical", "USL_DS", "GERBL2_2014_ComboC"): ("F", 4),
#     ("Historical", "SLR_US", "GERBL2_2014_ComboC"): ("G", 4),
#     ("Historical", "SLR_DS", "GERBL2_2014_ComboC"): ("H", 4),
#     ("RCP45", "LKO", "GERBL2_2014_ComboC"): ("D", 10),
#     ("RCP45", "USL_US", "GERBL2_2014_ComboC"): ("E", 10),
#     ("RCP45", "USL_DS", "GERBL2_2014_ComboC"): ("F", 10),
#     ("RCP45", "SLR_US", "GERBL2_2014_ComboC"): ("G", 10),
#     ("RCP45", "SLR_DS", "GERBL2_2014_ComboC"): ("H", 10),
#
#     ("Historical", "LKO", "GERBL2_2014_ComboD"): ("D", 5),
#     ("Historical", "USL_US", "GERBL2_2014_ComboD"): ("E", 5),
#     ("Historical", "USL_DS", "GERBL2_2014_ComboD"): ("F", 5),
#     ("Historical", "SLR_US", "GERBL2_2014_ComboD"): ("G", 5),
#     ("Historical", "SLR_DS", "GERBL2_2014_ComboD"): ("H", 5),
#     ("RCP45", "LKO", "GERBL2_2014_ComboD"): ("D", 11),
#     ("RCP45", "USL_US", "GERBL2_2014_ComboD"): ("E", 11),
#     ("RCP45", "USL_DS", "GERBL2_2014_ComboD"): ("F", 11),
#     ("RCP45", "SLR_US", "GERBL2_2014_ComboD"): ("G", 11),
#     ("RCP45", "SLR_DS", "GERBL2_2014_ComboD"): ("H", 11),
#
#     ("Historical", "LKO", "PreProject"): ("D", 6),
#     ("Historical", "USL_US", "PreProject"): ("E", 6),
#     ("Historical", "SLR_US", "PreProject"): ("G", 6),
#     ("Historical", "SLR_DS", "PreProject"): ("H", 6),
#     ("RCP45", "LKO", "PreProject"): ("D", 12),
#     ("RCP45", "USL_US", "PreProject"): ("E", 12),
#     ("RCP45", "SLR_US", "PreProject"): ("G", 12),
#     ("RCP45", "SLR_DS", "PreProject"): ("H", 12),
# }

list_section = ['LKO', 'USL_US', 'USL_DS', 'SLR_US', 'SLR_DS']
list_plan = ["GERBL2_2014", "GERBL2_2014_ComboC", "GERBL2_2014_ComboD", "PreProject"]
list_supply = ['Historical', 'STO330', 'RCP45']



cols = 'DEFGH'
start_line = 3
mapping_rules_3 = {}
for i in range(len(list_supply)):
    for j in range (len(list_plan)):
        for k in range(len(list_section)):
            line = start_line + (start_line * i*2) + j

            supply = list_supply[i]
            plan = list_plan[j]
            col = cols[k]
            sect = list_section[k]

            k = 0

            key = (supply, sect, plan)
            value = (col, line)
            mapping_rules_3[key] = value

list_section = ['LKO_CAN', 'LKO_US', 'USL_US_CAN', 'USL_US_US', 'USL_DS_CAN', 'USL_DS_US' , 'SLR_US_CAN', 'SLR_DS_CAN']
list_plan = ["GERBL2_2014", "GERBL2_2014_ComboC", "GERBL2_2014_ComboD", "PreProject"]
list_supply = ['Historical', 'STO330', 'RCP45']

mapping_country = {}

cols = 'DEFGHIJK'
start_line = 3
for i in range(len(list_supply)):
    for j in range (len(list_plan)):
        for k in range(len(list_section)):
            line = start_line + (start_line * i*2) + j

            supply = list_supply[i]
            plan = list_plan[j]
            col = cols[k]
            sect = list_section[k]

            key = (supply, sect, plan)
            value = (col, line)
            mapping_country[key] = value

            print(key, value)

#
# mapping_country =  {
#     ("Historical", "LKO_CAN", "GERBL2_2014"): ("E", 3),  # Maps to cell B5
#     ("Historical", "LKO_US", "GERBL2_2014"): ("F", 3),  # Maps to cell B5
#     ("Historical", "USL_US_CAN", "GERBL2_2014"): ("G", 3),  # Maps to cell B5
#     ("Historical", "USL_US_US", "GERBL2_2014"): ("H", 3),  # Maps to cell B5
#     ("Historical", "USL_DS_CAN", "GERBL2_2014"): ("I", 3),  # Maps to cell B5
#     ("Historical", "USL_DS_US", "GERBL2_2014"): ("J", 3),  # Maps to cell B5
#     ("Historical", "SLR_US_CAN", "GERBL2_2014"): ("K", 3),  # Maps to cell B5
#     ("Historical", "SLR_DS_CAN", "GERBL2_2014"): ("L", 3),  # Maps to cell B5
#
#     ("STO330", "LKO_CAN", "GERBL2_2014"): ("E", 8),  # Maps to cell B5
#     ("STO330", "LKO_US", "GERBL2_2014"): ("F", 8),  # Maps to cell B5
#     ("STO330", "USL_US_CAN", "GERBL2_2014"): ("G", 8),  # Maps to cell B5
#     ("STO330", "USL_US_US", "GERBL2_2014"): ("H", 8),  # Maps to cell B5
#     ("STO330", "USL_DS_CAN", "GERBL2_2014"): ("I", 8),  # Maps to cell B5
#     ("STO330", "USL_DS_US", "GERBL2_2014"): ("J", 8),  # Maps to cell B5
#     ("STO330", "SLR_US_CAN", "GERBL2_2014"): ("K", 8),  # Maps to cell B5
#     ("STO330", "SLR_DS_CAN", "GERBL2_2014"): ("L", 8),  # Maps to cell B5
#
#     ("RCP45", "LKO_CAN", "GERBL2_2014"): ("E", 13),  # Maps to cell B5
#     ("RCP45", "LKO_US", "GERBL2_2014"): ("F", 13),  # Maps to cell B5
#     ("RCP45", "USL_US_CAN", "GERBL2_2014"): ("G", 13),  # Maps to cell B5
#     ("RCP45", "USL_US_US", "GERBL2_2014"): ("H", 13),  # Maps to cell B5
#     ("RCP45", "USL_DS_CAN", "GERBL2_2014"): ("I", 13),  # Maps to cell B5
#     ("RCP45", "USL_DS_US", "GERBL2_2014"): ("J", 13),  # Maps to cell B5
#     ("RCP45", "SLR_US_CAN", "GERBL2_2014"): ("K", 13),  # Maps to cell B5
#     ("RCP45", "SLR_DS_CAN", "GERBL2_2014"): ("L", 13),  # Maps to cell B5
#
#     ("Historical", "LKO_CAN", "GERBL2_2014_ComboC"): ("E", 4),  # Maps to cell B5
#     ("Historical", "LKO_US", "GERBL2_2014_ComboC"): ("F", 4),  # Maps to cell B5
#     ("Historical", "USL_US_CAN", "GERBL2_2014_ComboC"): ("G", 4),  # Maps to cell B5
#     ("Historical", "USL_US_US", "GERBL2_2014_ComboC"): ("H", 4),  # Maps to cell B5
#     ("Historical", "USL_DS_CAN", "GERBL2_2014_ComboC"): ("I", 4),  # Maps to cell B5
#     ("Historical", "USL_DS_US", "GERBL2_2014_ComboC"): ("J", 4),  # Maps to cell B5
#     ("Historical", "SLR_US_CAN", "GERBL2_2014_ComboC"): ("K", 4),  # Maps to cell B5
#     ("Historical", "SLR_DS_CAN", "GERBL2_2014_ComboC"): ("L", 4),  # Maps to cell B5
#
#     ("STO330", "LKO_CAN", "GERBL2_2014_ComboC"): ("E", 9),  # Maps to cell B5
#     ("STO330", "LKO_US", "GERBL2_2014_ComboC"): ("F", 9),  # Maps to cell B5
#     ("STO330", "USL_US_CAN", "GERBL2_2014_ComboC"): ("G", 9),  # Maps to cell B5
#     ("STO330", "USL_US_US", "GERBL2_2014_ComboC"): ("H", 9),  # Maps to cell B5
#     ("STO330", "USL_DS_CAN", "GERBL2_2014_ComboC"): ("I", 9),  # Maps to cell B5
#     ("STO330", "USL_DS_US", "GERBL2_2014_ComboC"): ("J", 9),  # Maps to cell B5
#     ("STO330", "SLR_US_CAN", "GERBL2_2014_ComboC"): ("K", 9),  # Maps to cell B5
#     ("STO330", "SLR_DS_CAN", "GERBL2_2014_ComboC"): ("L", 9),  # Maps to cell B5
#
#     ("RCP45", "LKO_CAN", "GERBL2_2014_ComboC"): ("E", 14),  # Maps to cell B5
#     ("RCP45", "LKO_US", "GERBL2_2014_ComboC"): ("F", 14),  # Maps to cell B5
#     ("RCP45", "USL_US_CAN", "GERBL2_2014_ComboC"): ("G", 14),  # Maps to cell B5
#     ("RCP45", "USL_US_US", "GERBL2_2014_ComboC"): ("H", 14),  # Maps to cell B5
#     ("RCP45", "USL_DS_CAN", "GERBL2_2014_ComboC"): ("I", 14),  # Maps to cell B5
#     ("RCP45", "USL_DS_US", "GERBL2_2014_ComboC"): ("J", 14),  # Maps to cell B5
#     ("RCP45", "SLR_US_CAN", "GERBL2_2014_ComboC"): ("K", 14),  # Maps to cell B5
#     ("RCP45", "SLR_DS_CAN", "GERBL2_2014_ComboC"): ("L", 14),  # Maps to cell B5
#
#     ("Historical", "LKO_CAN", "PreProject"): ("E", 5),  # Maps to cell B5
#     ("Historical", "LKO_US", "PreProject"): ("F", 5),  # Maps to cell B5
#     ("Historical", "USL_US_CAN", "PreProject"): ("G", 5),  # Maps to cell B5
#     ("Historical", "USL_US_US", "PreProject"): ("H", 5),  # Maps to cell B5
#     #("Historical", "USL_DS_CAN", "PreProject"): ("I", 5),  # Maps to cell B5
#    # ("Historical", "USL_DS_US", "PreProject"): ("J", 5),  # Maps to cell B5
#     ("Historical", "SLR_US_CAN", "PreProject"): ("K", 5),  # Maps to cell B5
#     ("Historical", "SLR_DS_CAN", "PreProject"): ("L", 5),  # Maps to cell B5
#
#     ("STO330", "LKO_CAN", "PreProject"): ("E", 10),  # Maps to cell B5
#     ("STO330", "LKO_US", "PreProject"): ("F", 10),  # Maps to cell B5
#     ("STO330", "USL_US_CAN", "PreProject"): ("G", 10),  # Maps to cell B5
#     ("STO330", "USL_US_US", "PreProject"): ("H", 10),  # Maps to cell B5
#     #("STO330", "USL_DS_CAN", "PreProject"): ("I", 10),  # Maps to cell B5
#     #("STO330", "USL_DS_US", "PreProject"): ("J", 10),  # Maps to cell B5
#     ("STO330", "SLR_US_CAN", "PreProject"): ("K", 10),  # Maps to cell B5
#     ("STO330", "SLR_DS_CAN", "PreProject"): ("L", 10),  # Maps to cell B5
#
#     ("RCP45", "LKO_CAN", "PreProject"): ("E", 15),  # Maps to cell B5
#     ("RCP45", "LKO_US", "PreProject"): ("F", 15),  # Maps to cell B5
#     ("RCP45", "USL_US_CAN", "PreProject"): ("G", 15),  # Maps to cell B5
#     ("RCP45", "USL_US_US", "PreProject"): ("H", 15),  # Maps to cell B5
#     #("RCP45", "USL_DS_CAN", "PreProject"): ("I", 15),  # Maps to cell B5
#     #("RCP45", "USL_DS_US", "PreProject"): ("J", 15),  # Maps to cell B5
#     ("RCP45", "SLR_US_CAN", "PreProject"): ("K", 15),  # Maps to cell B5
#     ("RCP45", "SLR_DS_CAN", "PreProject"): ("L", 15),  # Maps to cell B5
#
# }
#
# mapping_country =  {
#     ("Historical", "LKO_CAN", "GERBL2_2014"): ("D", 3),  # Maps to cell B5
#     ("Historical", "LKO_US", "GERBL2_2014"): ("E", 3),  # Maps to cell B5
#     ("Historical", "USL_US_CAN", "GERBL2_2014"): ("F", 3),  # Maps to cell B5
#     ("Historical", "USL_US_US", "GERBL2_2014"): ("G", 3),  # Maps to cell B5
#     ("Historical", "USL_DS_CAN", "GERBL2_2014"): ("H", 3),  # Maps to cell B5
#     ("Historical", "USL_DS_US", "GERBL2_2014"): ("I", 3),  # Maps to cell B5
#     ("Historical", "SLR_US_CAN", "GERBL2_2014"): ("J", 3),  # Maps to cell B5
#     ("Historical", "SLR_DS_CAN", "GERBL2_2014"): ("K", 3),  # Maps to cell B5
#
#     ("RCP45", "LKO_CAN", "GERBL2_2014"): ("D", 9),  # Maps to cell B5
#     ("RCP45", "LKO_US", "GERBL2_2014"): ("E", 9),  # Maps to cell B5
#     ("RCP45", "USL_US_CAN", "GERBL2_2014"): ("F", 9),  # Maps to cell B5
#     ("RCP45", "USL_US_US", "GERBL2_2014"): ("G", 9),  # Maps to cell B5
#     ("RCP45", "USL_DS_CAN", "GERBL2_2014"): ("H", 9),  # Maps to cell B5
#     ("RCP45", "USL_DS_US", "GERBL2_2014"): ("I", 9),  # Maps to cell B5
#     ("RCP45", "SLR_US_CAN", "GERBL2_2014"): ("J", 9),  # Maps to cell B5
#     ("RCP45", "SLR_DS_CAN", "GERBL2_2014"): ("K", 9),  # Maps to cell B5
#
#     ("Historical", "LKO_CAN", "GERBL2_2014_ComboC"): ("D", 4),  # Maps to cell B5
#     ("Historical", "LKO_US", "GERBL2_2014_ComboC"): ("E", 4),  # Maps to cell B5
#     ("Historical", "USL_US_CAN", "GERBL2_2014_ComboC"): ("F", 4),  # Maps to cell B5
#     ("Historical", "USL_US_US", "GERBL2_2014_ComboC"): ("G", 4),  # Maps to cell B5
#     ("Historical", "USL_DS_CAN", "GERBL2_2014_ComboC"): ("H", 4),  # Maps to cell B5
#     ("Historical", "USL_DS_US", "GERBL2_2014_ComboC"): ("I", 4),  # Maps to cell B5
#     ("Historical", "SLR_US_CAN", "GERBL2_2014_ComboC"): ("J", 4),  # Maps to cell B5
#     ("Historical", "SLR_DS_CAN", "GERBL2_2014_ComboC"): ("K", 4),  # Maps to cell B5
#
#     ("RCP45", "LKO_CAN", "GERBL2_2014_ComboC"): ("D", 10),  # Maps to cell B5
#     ("RCP45", "LKO_US", "GERBL2_2014_ComboC"): ("E", 10),  # Maps to cell B5
#     ("RCP45", "USL_US_CAN", "GERBL2_2014_ComboC"): ("F", 10),  # Maps to cell B5
#     ("RCP45", "USL_US_US", "GERBL2_2014_ComboC"): ("G", 10),  # Maps to cell B5
#     ("RCP45", "USL_DS_CAN", "GERBL2_2014_ComboC"): ("H", 10),  # Maps to cell B5
#     ("RCP45", "USL_DS_US", "GERBL2_2014_ComboC"): ("I", 10),  # Maps to cell B5
#     ("RCP45", "SLR_US_CAN", "GERBL2_2014_ComboC"): ("J", 10),  # Maps to cell B5
#     ("RCP45", "SLR_DS_CAN", "GERBL2_2014_ComboC"): ("K", 10),  # Maps to cell B5
#
#     ("Historical", "LKO_CAN", "GERBL2_2014_ComboD"): ("D", 5),  # Maps to cell B5
#     ("Historical", "LKO_US", "GERBL2_2014_ComboD"): ("E", 5),  # Maps to cell B5
#     ("Historical", "USL_US_CAN", "GERBL2_2014_ComboD"): ("F", 5),  # Maps to cell B5
#     ("Historical", "USL_US_US", "GERBL2_2014_ComboD"): ("G", 5),  # Maps to cell B5
#     ("Historical", "USL_DS_CAN", "GERBL2_2014_ComboD"): ("H", 5),  # Maps to cell B5
#     ("Historical", "USL_DS_US", "GERBL2_2014_ComboD"): ("I", 5),  # Maps to cell B5
#     ("Historical", "SLR_US_CAN", "GERBL2_2014_ComboD"): ("J", 5),  # Maps to cell B5
#     ("Historical", "SLR_DS_CAN", "GERBL2_2014_ComboD"): ("K", 5),  # Maps to cell B5
#
#     ("RCP45", "LKO_CAN", "GERBL2_2014_ComboD"): ("D", 11),  # Maps to cell B5
#     ("RCP45", "LKO_US", "GERBL2_2014_ComboD"): ("E", 11),  # Maps to cell B5
#     ("RCP45", "USL_US_CAN", "GERBL2_2014_ComboD"): ("F", 11),  # Maps to cell B5
#     ("RCP45", "USL_US_US", "GERBL2_2014_ComboD"): ("G", 11),  # Maps to cell B5
#     ("RCP45", "USL_DS_CAN", "GERBL2_2014_ComboD"): ("H", 11),  # Maps to cell B5
#     ("RCP45", "USL_DS_US", "GERBL2_2014_ComboD"): ("I", 11),  # Maps to cell B5
#     ("RCP45", "SLR_US_CAN", "GERBL2_2014_ComboD"): ("J", 11),  # Maps to cell B5
#     ("RCP45", "SLR_DS_CAN", "GERBL2_2014_ComboD"): ("K", 11),  # Maps to cell B5
#
#
#     ("Historical", "LKO_CAN", "PreProject"): ("D", 6),  # Maps to cell B5
#     ("Historical", "LKO_US", "PreProject"): ("E", 6),  # Maps to cell B5
#     ("Historical", "USL_US_CAN", "PreProject"): ("F", 6),  # Maps to cell B5
#     ("Historical", "USL_US_US", "PreProject"): ("G", 6),  # Maps to cell B5
#     #("Historical", "USL_DS_CAN", "PreProject"): ("I", 5),  # Maps to cell B5
#    # ("Historical", "USL_DS_US", "PreProject"): ("J", 5),  # Maps to cell B5
#     ("Historical", "SLR_US_CAN", "PreProject"): ("J", 6),  # Maps to cell B5
#     ("Historical", "SLR_DS_CAN", "PreProject"): ("K", 6),  # Maps to cell B5
#
#     ("RCP45", "LKO_CAN", "PreProject"): ("D", 12),  # Maps to cell B5
#     ("RCP45", "LKO_US", "PreProject"): ("E", 12),  # Maps to cell B5
#     ("RCP45", "USL_US_CAN", "PreProject"): ("F", 12),  # Maps to cell B5
#     ("RCP45", "USL_US_US", "PreProject"): ("G", 12),  # Maps to cell B5
#     #("RCP45", "USL_DS_CAN", "PreProject"): ("I", 15),  # Maps to cell B5
#     #("RCP45", "USL_DS_US", "PreProject"): ("J", 15),  # Maps to cell B5
#     ("RCP45", "SLR_US_CAN", "PreProject"): ("J", 12),  # Maps to cell B5
#     ("RCP45", "SLR_DS_CAN", "PreProject"): ("K", 12),  # Maps to cell B5
#
# }

dict_pi_var = {
    'CHNI_2D': {'N breeding pairs':['EXCEED_COUNT', 'CRITICAL_DIFF']},

    'IXEX_RPI_2D': {'Weighted usable area (ha)': ['EXCEED_COUNT', 'CRITICAL_DIFF']},

    'SAUV_2D': {'Migration habitat (ha)': ['EXCEED_COUNT', 'CRITICAL_DIFF']},

    'BIRDS_2D': {'Abundance (n individuals)': ['EXCEED_COUNT', 'CRITICAL_DIFF']},

    'ONZI_OCCUPANCY_1D': {'Probability of muskrat lodge viability': ['EXCEED_COUNT', 'CRITICAL_DIFF'],
                'Percentage of the occupancy area in a cattail wetland': ['EXCEED_COUNT', 'CRITICAL_DIFF']},

    'TURTLE_1D': {'Blanding turtle winter survival probability': ['EXCEED_COUNT', 'CRITICAL_DIFF'],
                  'Snapping turtle winter survival probability': ['EXCEED_COUNT', 'CRITICAL_DIFF'],
                  'Painted turtle winter survival probability': ['EXCEED_COUNT', 'CRITICAL_DIFF']},

    'ZIPA_1D': {'Wildrice survival probability': ['EXCEED_COUNT', 'CRITICAL_DIFF']},

    'ERIW_MIN_1D': {'Exposed Riverbed Index': ['EXCEED_COUNT', 'CRITICAL_DIFF']},

    'CWRM_2D': {'Total Wetland area (ha)': ['MEAN', 'MEAN_DIRECTION'], 'Wet Meadow area (ha)': ['MEAN', 'MEAN_DIRECTION']},

    'IERM_2D': {'Total Wetland area (ha)': ['MEAN', 'MEAN_DIRECTION'], 'Wet Meadow area (ha)': ['MEAN', 'MEAN_DIRECTION']},

     'PIKE_2D': {'Habitat available for spawning and embryo-larval development (ha)': ['EXCEED_COUNT', 'CRITICAL_DIFF'],
                  'Habitat available for spawning (ha)': ['EXCEED_COUNT', 'CRITICAL_DIFF']},

    'AYL_2D': {'Average Yield Loss for all crops ($)': ['MEAN', 'MEAN_DIRECTION']},

    'ROADS_2D': {'Primary roads (Nb of QMs)': ['MEAN', 'MEAN_DIRECTION'],
                  'Secondary roads (Nb of QMs)': ['MEAN', 'MEAN_DIRECTION'],
                  'Tertiary roads (Nb of QMs)': ['MEAN', 'MEAN_DIRECTION'],
                  'All roads (Nb of QMs)': ['MEAN', 'MEAN_DIRECTION'],
                  'Primary roads (Length in m)': ['MEAN', 'MEAN_DIRECTION'],
                  'Secondary roads (Length in m)': ['MEAN', 'MEAN_DIRECTION'],
                  'Tertiary roads (Length in m)': ['MEAN', 'MEAN_DIRECTION'],
                  'All roads (Length in m)': ['MEAN', 'MEAN_DIRECTION']},

    'MFI_2D': {'Impacts during the navigation season': ['MEAN', 'MEAN_DIRECTION'],
                'Number of QMs with impacts': ['MEAN', 'MEAN_DIRECTION']},

    'NFB_2D': {'Accessory buildings (boolean)': ['MEAN', 'MEAN_DIRECTION'],
                'Accessory building (Nb of QMs)': ['MEAN', 'MEAN_DIRECTION'],
                'Strategic assets buildings (boolean)': ['MEAN', 'MEAN_DIRECTION'],
                'Strategic assets buildings (Nb of QMs)': ['MEAN', 'MEAN_DIRECTION'],
                'Non-residential (boolean)': ['MEAN', 'MEAN_DIRECTION'],
                'Non-residential (Nb of QMs)': ['MEAN', 'MEAN_DIRECTION'],
                'Residential (boolean)': ['MEAN', 'MEAN_DIRECTION'],
                'Residential (Nb of QMs)': ['MEAN', 'MEAN_DIRECTION'],
                'Total buildings (boolean)': ['MEAN', 'MEAN_DIRECTION'],
                'Total buildings (Nb of QMs)': ['MEAN', 'MEAN_DIRECTION']},

    'WASTE_WATER_2D': {'number of wastewater facilities exceeding the average discharge threshold': ['SUM', 'MEAN_DIRECTION'],
                       'weighted (duration, discharge) number of wastewater facilities impacted': ['SUM', 'MEAN_DIRECTION']},

    'WATER_INTAKES_2D': {'number of water intake facilities exceeding the nominal capacity threshold': ['SUM', 'MEAN_DIRECTION'],
                         'weighted (duration, capacity) number of intake facilities impacted': ['SUM', 'MEAN_DIRECTION']},
    'SHORE_PROT_STRUC': {'Wave overtopping (*10e-2)': ['VALUE', 'MEAN_DIRECTION'], 'Wave overflow (*10e-4)':['VALUE', 'MEAN_DIRECTION']}
    }



dict_pi_section = {
    'CHNI_2D': ['SLR_DS'],

    'IXEX_RPI_2D': ['SLR_DS'],

    'SAUV_2D': ['SLR_DS'],

    'BIRDS_2D': ['LKO'],

    'ONZI_OCCUPANCY_1D': ['LKO', 'USL_US', 'USL_DS', 'SLR_US', 'SLR_DS'],

    'TURTLE_1D': ['LKO', 'USL_US', 'USL_DS', 'SLR_US', 'SLR_DS'],

    'ZIPA_1D': ['LKO', 'USL_US', 'USL_DS', 'SLR_US', 'SLR_DS'],

    'ERIW_MIN_1D': ['LKO', 'USL_US', 'SLR_DS'],

    'CWRM_2D': ['LKO', 'USL_US', 'USL_DS'],

    'IERM_2D': ['SLR_DS'],

     'PIKE_2D': ['LKO', 'USL_US', 'USL_DS', 'SLR_DS'],

    'AYL_2D': ['SLR_DS'],

    'ROADS_2D': ['LKO', 'USL_US', 'USL_DS', 'SLR_DS'],

    'MFI_2D': ['LKO', 'USL_US', 'USL_DS', 'SLR_DS'],

    'NFB_2D': ['LKO_CAN', 'LKO_US', 'USL_US_CAN', 'USL_US_US', 'USL_DS_CAN', 'USL_DS_US', 'SLR_DS_CAN'],

    'WASTE_WATER_2D': ['LKO', 'USL_US', 'USL_DS', 'SLR_DS'],

    'WATER_INTAKES_2D': ['LKO', 'USL_US', 'USL_DS', 'SLR_DS'],

    'SHORE_PROT_STRUC': ['LKO']
    }

# 'WASTE_WATER_2D': ['number of wastewater facilities exceeding the average discharge threshold',
#                    'weighted (duration, discharge) number of wastewater facilities impacted']
# ,
#
# 'WATER_INTAKES_2D': ['number of water intake facilities exceeding the nominal capacity threshold',
#                      'weighted (duration, capacity) number of intake facilities impacted']
# }



#list_pis_critical_thresholds = ['BIRDS_2D', 'CHNI_2D', 'IXEX_RPI_2D', 'SAUV_2D', 'PIKE_2D', 'ERIW_MIN_1D', ]

list_pi_can_us = ['NFB_2D']

df_results = pd.read_csv(csv_path, sep=';', header=0)


output_path = os.path.join(outfolder, f'PIs_SUMMARY_RESULTS_FORMATTED_TABLE_20250212.xlsx')

if os.path.exists(output_path):
    os.remove(output_path)

workbook = Workbook()
workbook.save(output_path)

for pi, dict_var in dict_pi_var.items():
    sections = dict_pi_section[pi]
    for var, list_cols in dict_var.items():
        print(pi, var)

        value_col_ref = list_cols[0]
        signif_col = list_cols[1]
        if value_col_ref == 'MEAN':
            value_col_plan = 'MEAN (RESIDUALS)'
        elif value_col_ref == 'SUM':
            value_col_plan = 'DIFF SUM (2014BOC)'
        elif value_col_ref == 'VALUE':
            value_col_plan = 'DIFF VALUE'
        else:
            value_col_plan = value_col_ref
        df_pi = df_results[(df_results['PI_NAME'] == pi) &
                           (df_results['VARIABLE'] == var) &
                           (df_results['SECT_NAME'].isin(sections))]
        if pi in list_pi_can_us:
            sheet = 'CAN-US_final'
            map_rules = mapping_country
        else:
            sheet = 'Simple_final'
            map_rules = mapping_rules_3
        map_values_to_template(template_path, sheet, df_pi, output_path,
                               map_rules, value_col_ref, value_col_plan, signif_col)