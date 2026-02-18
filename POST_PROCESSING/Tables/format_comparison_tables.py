import pandas as pd
from openpyxl import load_workbook
from openpyxl import Workbook
import os
from pathlib import Path
import importlib
import sys
sys.path.append(str(Path(__file__).resolve().parent.parent.parent))
import POST_PROCESSING.ISEE.CFG_POST_PROCESS_ISEE as cfg
from copy import copy
from openpyxl.styles import PatternFill
import numpy as np

# After creating the table, you need to add a note for MFI and NFB
# about using GLRMM wl for SLR_US section

def copy_template_block(template_sheet, consolidated_sheet, start_row):
    """
    Copies the content and formatting of the template sheet to the consolidated sheet.

    Parameters:
    - template_sheet: the source template sheet.
    - consolidated_sheet: the destination sheet to copy to.
    - start_row: the starting row in the destination sheet.
    """
    for row_idx, row in enumerate(template_sheet.iter_rows(), start=start_row):
        for col_idx, cell in enumerate(row, start=1):
            new_cell = consolidated_sheet.cell(row=row_idx, column=col_idx, value=cell.value)
            if cell.has_style:
                new_cell._style = copy(cell._style)  # Copy cell style
            new_cell.number_format = cell.number_format
            new_cell.font = copy(cell.font)
            new_cell.fill = copy(cell.fill)
            new_cell.border = copy(cell.border)
            new_cell.alignment = copy(cell.alignment)

    # Copy merged cells
    for merged_range in template_sheet.merged_cells.ranges:
        consolidated_sheet.merge_cells(start_row=start_row + merged_range.min_row - 1,
                                       end_row=start_row + merged_range.max_row - 1,
                                       start_column=merged_range.min_col,
                                       end_column=merged_range.max_col)


def map_values_to_template(template_path, sheet, df_data, output_path, mapping_rules, value_col_ref, value_col_plan, signif_test_col, ref_plan='GERBL2_2014'):
    """
    Fills the Excel template with data from a CSV file and saves it.

    Parameters:
    - template_path: str, path to the Excel template file.
    - df_data: df, Dataframe containing the data.
    - output_path: str, path to the output Excel file.
    - mapping_rules: dict, rules defining which cells map to which conditions.
    - value_col: Column to write results into.
    """
    # Load the template and CSV data
    template = load_workbook(template_path)
    template_sheet = template[sheet]

    consolidated_workbook = load_workbook(output_path)
    consolidated_sheet = consolidated_workbook.active
    consolidated_sheet.title = "Consolidated Results"

    next_available_row = consolidated_sheet.max_row+1
    n_rows = template_sheet.max_row

    color_red = 'FFFF6666'
    color_green = 'FF90EE90'
    color_white = 'FFFFFFFF'

    # Get unique combinations of PI and Variable
    unique_pis = df_data[['PI_NAME', 'VARIABLE']].drop_duplicates()
    for _, row in unique_pis.iterrows():
        pi = row['PI_NAME']
        variable = row['VARIABLE']
        # print(pi, variable)


        # Filter data for the current PI/Variable
        filtered_data = df_data[(df_data['PI_NAME'] == pi) & (df_data['VARIABLE'] == variable)]
        # Copy the template block into the consolidated sheet
        copy_template_block(template_sheet, consolidated_sheet, next_available_row)

        consolidated_sheet.cell(row=next_available_row + 2, column=1, value=pi)
        consolidated_sheet.cell(row=next_available_row + 2, column=2, value=variable)


        # Update template cells based on rules
        for _, data_row in filtered_data.iterrows():
            section = data_row['SECT_NAME']
            plan = data_row['PLAN_NAME']
            if plan == ref_plan:
                value_col = value_col_ref
            else:
                value_col = value_col_plan

            supply_scenario = data_row['SUPPLY_SCEN']
            value = np.round(data_row[value_col], decimals=2)
            direction = data_row[signif_test_col]
            if direction == '-':
                color = color_red
            elif direction == '+':
                color = color_green
            else:
                color = color_white


            # Get the cell based on mapping rules
            # Get the cell based on mapping rules
            list_index = []
            # print(supply_scenario, section, plan)
            if (supply_scenario, section, plan) in mapping_rules:
                col_letter, relative_row = mapping_rules[(supply_scenario, section, plan)]
                col_index = ord(col_letter.upper()) - ord('A') + 1
                row_index = next_available_row + relative_row - 1
                # print(row_index, col_index)
                list_index.append((row_index, col_index))
                cell = consolidated_sheet.cell(row=row_index, column=col_index, value=value)
                cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")


        next_available_row += n_rows # +2 for spacing between blocks

    # Define the grey fill color
    grey_fill = PatternFill(start_color="FFB0B0B0", end_color="FFB0B0B0", fill_type="solid")

    # Fill empty cells with grey
    for row in consolidated_sheet.iter_rows(min_row=1, max_row=consolidated_sheet.max_row,
                                            min_col=4, max_col=consolidated_sheet.max_column):
        for cell in row:
            if cell.value is None:  # Check if the cell is empty
                cell.fill = grey_fill
    consolidated_workbook.save(output_path)
    #consolidated_workbook.close()


# Example usage
template_path = r"\\ECQCG1JWPASP002\projets$\GLAM\Dashboard\ISEE_Dash_portable\results_template_2.xlsx"
outfolder = os.path.join(cfg.POST_PROCESS_RES, r'PI_CSV_RESULTS_20260130')
csv_path = os.path.join(outfolder, f'PIs_SUMMARY_RESULTS_20260130.csv')
# Define mapping rules (example: map section, plan, supply_scenario to Excel cells)

# todo: modifier pour ajouter ComboA et ComboB
list_section = ['LKO', 'USL_US', 'USL_DS', 'SLR_US', 'SLR_DS']
list_plan = ["GERBL2_2014", "GERBL2_2014_ComboA", "GERBL2_2014_ComboB", "GERBL2_2014_ComboC", "GERBL2_2014_ComboD", "PreProject"]
list_supply = ['Historical', 'STO330', 'RCP45']

# todo: modifier pour ajouter ComboA et ComboB
# Mapping rules 3 = mapping rules pour l'ensemble des PIs sauf NFB (qui sépare par pays)
cols = 'DEFGH'
start_line = 3
mapping_rules_3 = {}
for i in range(len(list_supply)):
    for j in range (len(list_plan)):
        for k in range(len(list_section)):
            # line = start_line + (start_line * i*2) + j
            line = start_line + j + i*(len(list_plan)+2)

            supply = list_supply[i]
            plan = list_plan[j]
            col = cols[k]
            sect = list_section[k]

            k = 0

            key = (supply, sect, plan)
            value = (col, line)
            mapping_rules_3[key] = value

# print(mapping_rules_3)
list_section = ['LKO_CAN', 'LKO_US', 'USL_US_CAN', 'USL_US_US', 'USL_DS_CAN', 'USL_DS_US' , 'SLR_US_CAN', 'SLR_DS_CAN']
list_plan = ["GERBL2_2014", "GERBL2_2014_ComboA", "GERBL2_2014_ComboB", "GERBL2_2014_ComboC", "GERBL2_2014_ComboD", "PreProject"]
list_supply = ['Historical', 'STO330', 'RCP45']

# Mapping rules_country = mapping rules pour  NFB

mapping_country = {}

cols = 'DEFGHIJK'
start_line = 3
for i in range(len(list_supply)):
    for j in range (len(list_plan)):
        for k in range(len(list_section)):
            # line = start_line + (start_line * i*2) + j
            line = start_line + j + i*(len(list_plan)+2)

            supply = list_supply[i]
            plan = list_plan[j]
            col = cols[k]
            sect = list_section[k]

            key = (supply, sect, plan)
            value = (col, line)
            mapping_country[key] = value

            # print(key, value)

dict_pi_var = {
    'AYL_2D': {'Average Yield Loss for all crops ($)': ['MEAN', 'MEAN_DIRECTION']},
    'BIRDS_2D': {'Abundance (n individuals)': ['EXCEED_COUNT', 'CRITICAL_DIFF']},
    'CHNI_2D': {'N breeding pairs':['EXCEED_COUNT', 'CRITICAL_DIFF']},
    'CWRM_2D': {'Total Wetland area (ha)': ['MEAN', 'MEAN_DIRECTION'], 'Wet Meadow area (ha)': ['MEAN', 'MEAN_DIRECTION']},
    'ERIW_MIN_1D': {'Exposed Riverbed Index': ['EXCEED_COUNT', 'CRITICAL_DIFF']},
    'IERM_2D': {'Total Wetland area (ha)': ['MEAN', 'MEAN_DIRECTION'], 'Wet Meadow area (ha)': ['MEAN', 'MEAN_DIRECTION']},
    'IXEX_RPI_2D': {'Weighted usable area (ha)': ['EXCEED_COUNT', 'CRITICAL_DIFF']},
    'MFI_2D': {'Impacts during the navigation season': ['MEAN', 'MEAN_DIRECTION'],'Number of QMs with impacts': ['MEAN', 'MEAN_DIRECTION']},

    'NFB_2D': {'Accessory buildings (boolean)': ['MEAN', 'MEAN_DIRECTION'],'Accessory building (Nb of QMs)': ['MEAN', 'MEAN_DIRECTION'],
                'Strategic assets buildings (boolean)': ['MEAN', 'MEAN_DIRECTION'],'Strategic assets buildings (Nb of QMs)': ['MEAN', 'MEAN_DIRECTION'],
                'Non-residential (boolean)': ['MEAN', 'MEAN_DIRECTION'],'Non-residential (Nb of QMs)': ['MEAN', 'MEAN_DIRECTION'],
                'Residential (boolean)': ['MEAN', 'MEAN_DIRECTION'],'Residential (Nb of QMs)': ['MEAN', 'MEAN_DIRECTION'],
                'Total buildings (boolean)': ['MEAN', 'MEAN_DIRECTION'],'Total buildings (Nb of QMs)': ['MEAN', 'MEAN_DIRECTION']},
    'ONZI_OCCUPANCY_1D': {'Probability of muskrat lodge viability': ['EXCEED_COUNT', 'CRITICAL_DIFF'],
                          'Percentage of the occupancy area in a cattail wetland': ['EXCEED_COUNT', 'CRITICAL_DIFF']},
    'PIKE_2D': {'Habitat available for spawning and embryo-larval development (ha)': ['EXCEED_COUNT', 'CRITICAL_DIFF'],
                  'Habitat available for spawning (ha)': ['EXCEED_COUNT', 'CRITICAL_DIFF']},
    'ROADS_2D': {'Primary roads (Nb of QMs)': ['MEAN', 'MEAN_DIRECTION'],'Secondary roads (Nb of QMs)': ['MEAN', 'MEAN_DIRECTION'],
                  'Tertiary roads (Nb of QMs)': ['MEAN', 'MEAN_DIRECTION'],'All roads (Nb of QMs)': ['MEAN', 'MEAN_DIRECTION'],
                  'Primary roads (Length in m)': ['MEAN', 'MEAN_DIRECTION'],'Secondary roads (Length in m)': ['MEAN', 'MEAN_DIRECTION'],
                  'Tertiary roads (Length in m)': ['MEAN', 'MEAN_DIRECTION'],'All roads (Length in m)': ['MEAN', 'MEAN_DIRECTION']},
    'SAUV_2D': {'Migration habitat (ha)': ['EXCEED_COUNT', 'CRITICAL_DIFF']},
    'SHORE_PROT_STRUC': {'Wave overtopping (*10e-2)': ['VALUE', 'MEAN_DIRECTION'], 'Wave overflow (*10e-4)':['VALUE', 'MEAN_DIRECTION']},
    'TURTLE_1D': {'Blanding turtle winter survival probability': ['EXCEED_COUNT', 'CRITICAL_DIFF'],
                  'Snapping turtle winter survival probability': ['EXCEED_COUNT', 'CRITICAL_DIFF'],
                  'Painted turtle winter survival probability': ['EXCEED_COUNT', 'CRITICAL_DIFF']},
    'WASTE_WATER_2D': {'number of wastewater facilities exceeding the average discharge threshold': ['SUM', 'MEAN_DIRECTION'],
                       'weighted (duration, discharge) number of wastewater facilities impacted': ['SUM', 'MEAN_DIRECTION']},
    'WATER_INTAKES_2D': {'number of water intake facilities exceeding the nominal capacity threshold': ['SUM', 'MEAN_DIRECTION'],
                         'weighted (duration, capacity) number of intake facilities impacted': ['SUM', 'MEAN_DIRECTION']},
    'ZIPA_1D': {'Wildrice survival probability': ['EXCEED_COUNT', 'CRITICAL_DIFF']}}

dict_pi_section = {
    'AYL_2D': ['SLR_DS'],
    'BIRDS_2D': ['LKO'],
    'CHNI_2D': ['SLR_DS'],
    'CWRM_2D': ['LKO', 'USL_US', 'USL_DS'],
    'ERIW_MIN_1D': ['LKO', 'USL_US', 'SLR_DS'],
    'IERM_2D': ['SLR_DS'],
    'IXEX_RPI_2D': ['SLR_DS'],
    'MFI_2D': ['LKO', 'USL_US', 'USL_DS', 'SLR_DS'],
    'NFB_2D': ['LKO_CAN', 'LKO_US', 'USL_US_CAN', 'USL_US_US', 'USL_DS_CAN', 'USL_DS_US', 'SLR_DS_CAN'],
    'ONZI_OCCUPANCY_1D': ['LKO', 'USL_US', 'USL_DS', 'SLR_US', 'SLR_DS'],
    'PIKE_2D': ['LKO', 'USL_US', 'USL_DS', 'SLR_DS'],
    'ROADS_2D': ['LKO', 'USL_US', 'USL_DS', 'SLR_DS'],
    'SAUV_2D': ['SLR_DS'],
    'SHORE_PROT_STRUC': ['LKO'],
    'TURTLE_1D': ['LKO', 'USL_US', 'USL_DS', 'SLR_US', 'SLR_DS'],
    'WASTE_WATER_2D': ['LKO', 'USL_US', 'USL_DS', 'SLR_DS'],
    'WATER_INTAKES_2D': ['LKO', 'USL_US', 'USL_DS', 'SLR_DS'],
    'ZIPA_1D': ['LKO', 'USL_US', 'USL_DS', 'SLR_US', 'SLR_DS']}


list_pi_can_us = ['NFB_2D']

df_results = pd.read_csv(csv_path, sep=';', header=0)

output_path = os.path.join(outfolder, f'PIs_SUMMARY_RESULTS_FORMATTED_TABLE_20260130.xlsx')

if os.path.exists(output_path):
    os.remove(output_path)

workbook = Workbook()
workbook.save(output_path)

for pi, dict_var in dict_pi_var.items():
    sections = dict_pi_section[pi]
    for var, list_cols in dict_var.items():
        print(pi, var, sections)

        value_col_ref = list_cols[0]
        signif_col = list_cols[1]
        if value_col_ref == 'MEAN':
            value_col_plan = 'MEAN (RESIDUALS)'
        elif value_col_ref == 'SUM':
            value_col_plan = 'DIFF SUM (2014BOC)'
        elif value_col_ref == 'VALUE':
            value_col_plan = 'DIFF VALUE'
        else:
            value_col_plan = value_col_ref

        df_pi = df_results[(df_results['PI_NAME'] == pi) &
                           (df_results['VARIABLE'] == var) &
                           (df_results['SECT_NAME'].isin(sections))]

        if pi in list_pi_can_us:
            sheet = 'CAN-US_final'
            map_rules = mapping_country
        else:
            sheet = 'Simple_final'
            map_rules = mapping_rules_3
        map_values_to_template(template_path, sheet, df_pi, output_path,
                               map_rules, value_col_ref, value_col_plan, signif_col)